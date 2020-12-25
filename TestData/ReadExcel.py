# Import Package
import openpyxl

# Load Workbook

vk = openpyxl.load_workbook("C:\\Users\\225027\\PycharmProjects\\pythonProject\\JD_TestData.xlsx")


def fetch_number_of_rows(Sheetname):
    sh = vk[Sheetname]
    return sh.max_row


def fetch_cell_data(Sheetname, row, cell):
    sh = vk[Sheetname]
    cell = sh.cell(row, cell)
    return cell.value


print(fetch_number_of_rows("TestData"))
print(fetch_cell_data("TestData", 2, 1))
print(fetch_cell_data("TestData", 2, 2))

# print(vk.sheetnames)
# print("Active sheet is "+ vk.active.title)

# sh = vk["TestData"]
# print(sh.title)
# print(sh.max_row)
